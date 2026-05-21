"""
Router para Acompanhamento de Frequência - Bolsa Família.
"""

from fastapi import APIRouter, HTTPException, Request, Query
from fastapi.responses import StreamingResponse
from typing import Optional, List
from datetime import datetime, timezone, date, timedelta
import logging
import io
import calendar

from auth_middleware import AuthMiddleware
from pdf_cache import get_mantenedora_cached
from pdf.utils import format_date_pt
from services.attendance_utils import (
    compute_monthly_valid_absences,
    fetch_medical_days_for_students,
)
from services.bf_reason_suggestion import suggest_reason_for_month
from services.bf_network_stats import (
    compute_network_stats,
    list_followup_cases,
    STATS_VERSION as NETWORK_STATS_VERSION,
)
from services.bf_legacy_migration import (
    preview_legacy_migration,
    apply_legacy_migration,
)

logger = logging.getLogger(__name__)

router = APIRouter(tags=["Bolsa Família"])

MESES_PT = {
    1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril',
    5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto',
    9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
}


def setup_router(db, **kwargs):

    async def _calc_monthly_school_days(academic_year):
        """Calcula dias letivos por mês usando a mesma lógica do módulo de frequência (attendance.py)."""
        # Buscar calendário letivo (campo correto: ano_letivo, global: school_id=None)
        calendario = await db.calendario_letivo.find_one(
            {"ano_letivo": academic_year, "school_id": None}, {"_id": 0}
        )
        if not calendario:
            calendario = await db.calendario_letivo.find_one(
                {"ano_letivo": academic_year}, {"_id": 0}
            )

        if not calendario:
            # Sem calendário, retorna zeros
            return {m: 0 for m in range(1, 13)}

        # Buscar eventos do calendário
        eventos_nao_letivos = ['feriado_nacional', 'feriado_estadual', 'feriado_municipal', 'recesso_escolar']
        events = await db.calendar_events.find(
            {"academic_year": academic_year}, {"_id": 0}
        ).to_list(1000)

        datas_nao_letivas = set()
        datas_sabados_letivos = set()
        for event in events:
            event_type = event.get('event_type', '')
            start_date_str = event.get('start_date')
            end_date_str = event.get('end_date') or start_date_str
            if not start_date_str:
                continue
            try:
                start_d = datetime.strptime(start_date_str[:10], '%Y-%m-%d').date()
                end_d = datetime.strptime(end_date_str[:10], '%Y-%m-%d').date()
                current = start_d
                while current <= end_d:
                    if event_type in eventos_nao_letivos:
                        datas_nao_letivas.add(current)
                    elif event_type == 'sabado_letivo':
                        datas_sabados_letivos.add(current)
                    elif event.get('is_school_day', False) and current.weekday() == 5:
                        datas_sabados_letivos.add(current)
                    current += timedelta(days=1)
            except (ValueError, TypeError):
                continue

        def _contar_dias_periodo(inicio_str, fim_str):
            if not inicio_str or not fim_str:
                return 0
            try:
                inicio = datetime.strptime(str(inicio_str)[:10], '%Y-%m-%d').date()
                fim = datetime.strptime(str(fim_str)[:10], '%Y-%m-%d').date()
            except (ValueError, TypeError):
                return 0
            dias = 0
            current = inicio
            while current <= fim:
                if current in datas_sabados_letivos:
                    dias += 1
                elif current.weekday() < 5:
                    if current not in datas_nao_letivas:
                        dias += 1
                current += timedelta(days=1)
            return dias

        # Coletar todas as datas letivas dos 4 bimestres
        def _coletar_datas_periodo(inicio_str, fim_str):
            datas = []
            if not inicio_str or not fim_str:
                return datas
            try:
                inicio = datetime.strptime(str(inicio_str)[:10], '%Y-%m-%d').date()
                fim = datetime.strptime(str(fim_str)[:10], '%Y-%m-%d').date()
            except (ValueError, TypeError):
                return datas
            current = inicio
            while current <= fim:
                is_school = False
                if current in datas_sabados_letivos:
                    is_school = True
                elif current.weekday() < 5 and current not in datas_nao_letivas:
                    is_school = True
                if is_school:
                    datas.append(current)
                current += timedelta(days=1)
            return datas

        # Agregar dias letivos por mês a partir das datas dos bimestres
        monthly = {m: 0 for m in range(1, 13)}
        bimestres = [
            (calendario.get('bimestre_1_inicio'), calendario.get('bimestre_1_fim')),
            (calendario.get('bimestre_2_inicio'), calendario.get('bimestre_2_fim')),
            (calendario.get('bimestre_3_inicio'), calendario.get('bimestre_3_fim')),
            (calendario.get('bimestre_4_inicio'), calendario.get('bimestre_4_fim')),
        ]
        for b_inicio, b_fim in bimestres:
            datas = _coletar_datas_periodo(b_inicio, b_fim)
            for dt in datas:
                monthly[dt.month] += 1

        return monthly

    async def _calc_student_monthly_attendance(student_id, academic_year, months_range):
        """Calcula a frequência real mensal de um aluno com base nos registros de presença."""
        attendance_records = await db.attendance.find(
            {"student_id": student_id, "academic_year": academic_year},
            {"_id": 0, "date": 1, "status": 1, "records": 1}
        ).to_list(10000)

        monthly_presence = {}
        monthly_total = {}

        for rec in attendance_records:
            rec_date = rec.get("date", "")
            if not rec_date:
                continue
            try:
                dt = datetime.strptime(rec_date[:10], "%Y-%m-%d")
                m = dt.month
            except:
                continue

            if m not in months_range:
                continue

            # Check individual records or status
            records = rec.get("records", [])
            if records:
                for r in records:
                    st = r.get("status", "")
                    if st:
                        monthly_total[m] = monthly_total.get(m, 0) + 1
                        if st in ("present", "presente", "P"):
                            monthly_presence[m] = monthly_presence.get(m, 0) + 1
            else:
                status = rec.get("status", "")
                if status:
                    monthly_total[m] = monthly_total.get(m, 0) + 1
                    if status in ("present", "presente", "P"):
                        monthly_presence[m] = monthly_presence.get(m, 0) + 1

        return monthly_presence, monthly_total

    EDIT_ROLES = ['super_admin', 'admin', 'admin_teste', 'secretario', 'gerente']
    VIEW_ROLES = ['super_admin', 'admin', 'admin_teste', 'secretario', 'semed3', 'diretor', 'ass_social_2', 'gerente']
    # Roles autorizadas a consolidar TODAS AS ESCOLAS (visão de rede)
    ALL_SCHOOLS_ROLES = ['super_admin', 'admin', 'gerente', 'semed3']

    @router.get("/bolsa-familia/reason-groups")
    async def list_reason_groups(request: Request, mec_version: Optional[str] = "4.2"):
        """Lista grupos oficiais MEC de motivos de baixa frequência."""
        await AuthMiddleware.require_permission(
            db, 'nav-bolsa-familia-button', VIEW_ROLES
        )(request)
        query = {"active": True}
        if mec_version:
            query["mec_version"] = mec_version
        groups = await db.attendance_frequency_reason_groups.find(
            query, {"_id": 0}
        ).sort("sort_order", 1).to_list(1000)
        return {"groups": groups, "mec_version": mec_version, "total": len(groups)}

    @router.get("/bolsa-familia/reasons")
    async def list_reasons(
        request: Request,
        group_id: Optional[str] = None,
        mec_version: Optional[str] = "4.2",
        include_legacy: bool = False,
    ):
        """Lista submotivos MEC. Opcionalmente filtra por grupo.

        Resposta enxuta — sem agrupamento. Frontend faz `group by group_id`.
        """
        await AuthMiddleware.require_permission(
            db, 'nav-bolsa-familia-button', VIEW_ROLES
        )(request)
        query = {"active": True}
        if mec_version:
            query["mec_version"] = mec_version
        if group_id:
            query["group_id"] = group_id
        if not include_legacy:
            query["legacy"] = {"$ne": True}
        reasons = await db.attendance_frequency_reasons.find(
            query, {"_id": 0}
        ).sort([("mec_group_code", 1), ("mec_subcode", 1)]).to_list(5000)
        return {"reasons": reasons, "mec_version": mec_version, "total": len(reasons)}

    @router.get("/bolsa-familia/reasons/grouped")
    async def list_reasons_grouped(
        request: Request,
        mec_version: Optional[str] = "4.2",
        include_legacy: bool = False,
    ):
        """Lista submotivos já agrupados (estrutura ideal para Combobox).

        Shape: `{groups: [{group_id, mec_code, name, category, sort_order,
        reasons: [{id, mec_subcode, name, severity_level, requires_followup}]}]}`
        """
        await AuthMiddleware.require_permission(
            db, 'nav-bolsa-familia-button', VIEW_ROLES
        )(request)

        groups_q = {"active": True}
        if mec_version:
            groups_q["mec_version"] = mec_version
        groups = await db.attendance_frequency_reason_groups.find(
            groups_q, {"_id": 0}
        ).sort("sort_order", 1).to_list(1000)

        reasons_q = {"active": True}
        if mec_version:
            reasons_q["mec_version"] = mec_version
        if not include_legacy:
            reasons_q["legacy"] = {"$ne": True}
        reasons = await db.attendance_frequency_reasons.find(
            reasons_q, {"_id": 0}
        ).sort([("mec_group_code", 1), ("mec_subcode", 1)]).to_list(5000)

        # Agrupa por group_id
        by_group: dict = {}
        for r in reasons:
            by_group.setdefault(r["group_id"], []).append(r)

        result = []
        for g in groups:
            result.append({
                "group_id": g["id"],
                "mec_code": g["mec_code"],
                "name": g["name"],
                "category": g.get("category"),
                "sort_order": g.get("sort_order"),
                "reasons": by_group.get(g["id"], []),
            })
        return {"groups": result, "mec_version": mec_version}

    @router.get("/bolsa-familia/suggest-reason")
    async def suggest_reason(
        request: Request,
        student_id: str = Query(...),
        school_id: str = Query(...),
        month: int = Query(..., ge=1, le=12),
        academic_year: Optional[int] = None,
    ):
        """Suggestion Engine determinística (Fase 2 — Fev/2026).

        Retorna sugestão de motivo MEC para um par (aluno, mês) baseada em
        regras explícitas, auditáveis, sem IA. Frontend pode mostrar a
        sugestão como hint visual; operador decide se aceita.

        Resposta:
        ```
        {
          "engine_version": "1.0",
          "suggested_reason_id": str | None,
          "suggested_reason_subcode": str | None,
          "confidence": float (0..1),
          "rules_triggered": [{code, value, threshold, ...}],
          "requires_followup_flag": bool,
          "human_explanation": str,
          "should_show_suggestion": bool,
          "metrics": {
            "school_days": N, "valid_absences": N, "medical_days_count": N,
            "frequency_percentage": float | None
          }
        }
        ```
        """
        await AuthMiddleware.require_permission(
            db, 'nav-bolsa-familia-button', VIEW_ROLES
        )(request)
        if not academic_year:
            academic_year = datetime.now().year

        # 1) Calcular métricas do mês usando engine canônica
        monthly_school_days = await _calc_monthly_school_days(academic_year)
        school_days = monthly_school_days.get(month, 0)

        # Busca registros de attendance do aluno no ano (escopado por escola).
        # Filtro adicional client-side por month — pequeno volume.
        all_attendance = await db.attendance.find(
            {"academic_year": academic_year},
            {"_id": 0, "date": 1, "records": 1},
        ).to_list(50000)

        medical_days_by_student = await fetch_medical_days_for_students(
            db, [student_id], academic_year
        )
        absences_map = compute_monthly_valid_absences(
            all_attendance, medical_days_by_student, {student_id}
        )
        valid_absences = (absences_map.get(student_id) or {}).get(month, 0)
        medical_set = medical_days_by_student.get(student_id) or set()
        medical_days_count = sum(
            1 for d in medical_set if len(d) == 10 and d[5:7] == f"{month:02d}"
        )

        frequency_percentage: Optional[float] = None
        if school_days > 0:
            frequency_percentage = round(
                ((school_days - valid_absences) * 100) / school_days, 1
            )

        # 2) Carrega reasons MEC para resolver subcode → id
        reasons_cursor = db.attendance_frequency_reasons.find(
            {"active": True, "mec_version": "4.2"},
            {"_id": 0, "id": 1, "mec_subcode": 1, "group_id": 1, "severity_level": 1,
             "requires_followup": 1, "name": 1},
        )
        reasons_by_subcode: dict = {}
        async for r in reasons_cursor:
            reasons_by_subcode[r["mec_subcode"]] = r

        # 3) Resolve motivo atual (se houver) para R3 (severity flag)
        tracking_doc = await db.bolsa_familia_tracking.find_one(
            {
                "student_id": student_id,
                "school_id": school_id,
                "month": str(month),
                "academic_year": academic_year,
            },
            {"_id": 0, "reason_id": 1},
        )
        current_reason: Optional[dict] = None
        if tracking_doc and tracking_doc.get("reason_id"):
            current_reason = await db.attendance_frequency_reasons.find_one(
                {"id": tracking_doc["reason_id"]},
                {"_id": 0, "id": 1, "severity_level": 1, "requires_followup": 1,
                 "name": 1, "mec_subcode": 1},
            )

        # 4) Roda engine pura (lógica isolada em services/bf_reason_suggestion.py)
        result = suggest_reason_for_month(
            medical_days_count=medical_days_count,
            valid_absences=valid_absences,
            school_days=school_days,
            frequency_percentage=frequency_percentage,
            reasons_by_subcode=reasons_by_subcode,
            current_reason=current_reason,
        )
        result["metrics"] = {
            "school_days": school_days,
            "valid_absences": valid_absences,
            "medical_days_count": medical_days_count,
            "frequency_percentage": frequency_percentage,
        }
        return result

    # ------------------------------------------------------------------
    # Fase 3A — Agregados institucionais (cache in-process TTL 5min).
    # Owner spec: backend PRIMEIRO, dashboard depois. Pipeline `$facet`
    # única — sem múltiplas queries dispersas.
    # ------------------------------------------------------------------
    _STATS_CACHE_TTL_SECONDS = 300
    _stats_cache: dict = {}

    def _cache_get(key: str):
        entry = _stats_cache.get(key)
        if not entry:
            return None
        if (datetime.now(timezone.utc) - entry["at"]).total_seconds() > _STATS_CACHE_TTL_SECONDS:
            _stats_cache.pop(key, None)
            return None
        return entry["data"]

    def _cache_set(key: str, data):
        _stats_cache[key] = {"at": datetime.now(timezone.utc), "data": data}

    def _cache_invalidate_all():
        """Invalida todo o cache de stats — chamado quando trackings mudam."""
        _stats_cache.clear()

    @router.get("/bolsa-familia/stats/network")
    async def network_stats(
        request: Request,
        academic_year: Optional[int] = None,
        mec_version: str = "4.2",
        force_refresh: bool = False,
    ):
        """Agregados institucionais da rede — pronto para Dashboard Busca Ativa.

        Resposta tem `stats_version` para evolução segura. Cache in-process
        com TTL de 5 min (evita explosão por polling de dashboard).
        """
        await AuthMiddleware.require_permission(
            db, 'nav-bolsa-familia-button', VIEW_ROLES
        )(request)

        cache_key = f"network|{academic_year}|{mec_version}"
        if not force_refresh:
            cached = _cache_get(cache_key)
            if cached is not None:
                return {**cached, "cached": True}

        data = await compute_network_stats(
            db, academic_year=academic_year, mec_version=mec_version
        )
        _cache_set(cache_key, data)
        return {**data, "cached": False}

    @router.get("/bolsa-familia/stats/network/followup")
    async def network_followup(
        request: Request,
        academic_year: Optional[int] = None,
        mec_version: str = "4.2",
        severity_min: int = Query(5, ge=1, le=5),
        limit: int = Query(200, ge=1, le=1000),
        category: Optional[str] = None,
        school_id: Optional[str] = None,
    ):
        """Casos prioritários para Busca Ativa: severity >= N ou
        requires_followup=True. Lista limitada (`limit`), nunca dump completo.

        Filtros Fase 3B: `category` (HEALTH, VIOLENCE, ACCESS, …) e `school_id`.
        """
        await AuthMiddleware.require_permission(
            db, 'nav-bolsa-familia-button', VIEW_ROLES
        )(request)
        return await list_followup_cases(
            db,
            academic_year=academic_year,
            mec_version=mec_version,
            severity_min=severity_min,
            limit=limit,
            category=category,
            school_id=school_id,
        )

    @router.post("/bolsa-familia/stats/network/snapshot")
    async def take_network_snapshot(
        request: Request,
        academic_year: Optional[int] = None,
        mec_version: str = "4.2",
    ):
        """Persiste snapshot diário dos agregados — histórico longitudinal.

        Idempotente por (snapshot_date, academic_year, mec_version). Pensado
        para ser invocado por job cron noturno externo:
            POST /api/bolsa-familia/stats/network/snapshot?academic_year=2026
        """
        await AuthMiddleware.require_permission(
            db, 'nav-bolsa-familia-button', EDIT_ROLES
        )(request)
        if not academic_year:
            academic_year = datetime.now().year
        snap = await compute_network_stats(
            db, academic_year=academic_year, mec_version=mec_version
        )
        snapshot_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        doc = {
            "snapshot_date": snapshot_date,
            "stats_version": snap["stats_version"],
            "scope": snap["scope"],
            "payload": snap,
            "saved_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.bf_network_stats_snapshots.update_one(
            {
                "snapshot_date": snapshot_date,
                "scope.academic_year": academic_year,
                "scope.mec_version": mec_version,
            },
            {"$set": doc},
            upsert=True,
        )
        return {"saved": True, "snapshot_date": snapshot_date,
                "scope": snap["scope"]}

    @router.get("/bolsa-familia/stats/snapshots")
    async def list_snapshots(
        request: Request,
        academic_year: Optional[int] = None,
        mec_version: str = "4.2",
        from_date: Optional[str] = None,
        to_date: Optional[str] = None,
        limit: int = Query(180, ge=1, le=730),
    ):
        """Histórico de snapshots para gráficos de evolução temporal."""
        await AuthMiddleware.require_permission(
            db, 'nav-bolsa-familia-button', VIEW_ROLES
        )(request)
        query: dict = {"scope.mec_version": mec_version}
        if academic_year is not None:
            query["scope.academic_year"] = academic_year
        if from_date or to_date:
            date_filter: dict = {}
            if from_date:
                date_filter["$gte"] = from_date
            if to_date:
                date_filter["$lte"] = to_date
            query["snapshot_date"] = date_filter
        snaps = await db.bf_network_stats_snapshots.find(
            query,
            {"_id": 0, "snapshot_date": 1, "stats_version": 1, "scope": 1,
             "payload.total_with_reason": 1, "payload.by_category": 1,
             "payload.by_severity": 1, "payload.requires_followup": 1,
             "payload.severity_5_plus": 1},
        ).sort("snapshot_date", -1).to_list(limit)
        series = [
            {
                "snapshot_date": s["snapshot_date"],
                "stats_version": s["stats_version"],
                "total_with_reason": (s.get("payload") or {}).get("total_with_reason", 0),
                "by_category": (s.get("payload") or {}).get("by_category", {}),
                "by_severity": (s.get("payload") or {}).get("by_severity", {}),
                "requires_followup": (s.get("payload") or {}).get("requires_followup", 0),
                "severity_5_plus": (s.get("payload") or {}).get("severity_5_plus", 0),
            }
            for s in snaps
        ]
        return {"stats_version": NETWORK_STATS_VERSION, "total": len(series),
                "series": series}

    @router.get("/bolsa-familia/migrate-legacy/preview")
    async def preview_legacy(
        request: Request,
        academic_year: Optional[int] = None,
        confidence_min: float = Query(0.0, ge=0.0, le=1.0),
        limit: int = Query(5000, ge=1, le=20000),
    ):
        """Preview da migração legacy → MEC (Fev/2026 Fase 3C).

        Analisa todos os trackings com `motive_legacy` e propõe `reason_id`
        baseado em regex/keyword. NÃO persiste nada. Owner spec: gestor
        revisa antes de aplicar.
        """
        await AuthMiddleware.require_permission(
            db, 'nav-bolsa-familia-button', VIEW_ROLES
        )(request)
        return await preview_legacy_migration(
            db,
            academic_year=academic_year,
            confidence_min=confidence_min,
            limit=limit,
        )

    @router.post("/bolsa-familia/migrate-legacy/apply")
    async def apply_legacy(
        request: Request,
        academic_year: Optional[int] = None,
        confidence_min: float = Query(0.7, ge=0.0, le=1.0),
        include_fallback: bool = Query(False),
    ):
        """Aplica migração legacy → MEC. Apenas EDIT_ROLES.

        - `confidence_min`: só migra matches com confidence >= esse valor.
        - `include_fallback=true`: força marcar não classificados como `24z`.
        - Idempotente — documentos já com `reason_id` não são alterados.
        - Cada documento migrado ganha `legacy_migration: {audit metadata}`.
        """
        current_user = await AuthMiddleware.require_permission(
            db, 'nav-bolsa-familia-button', EDIT_ROLES
        )(request)
        result = await apply_legacy_migration(
            db,
            academic_year=academic_year,
            confidence_min=confidence_min,
            include_fallback=include_fallback,
            user_id=current_user.get("id"),
        )
        if result.get("migrated", 0) > 0:
            _cache_invalidate_all()
        return result



    @router.get("/bolsa-familia/stats/network/followup/export")
    async def network_followup_export(
        request: Request,
        academic_year: Optional[int] = None,
        mec_version: str = "4.2",
        severity_min: int = Query(1, ge=1, le=5),
        limit: int = Query(1000, ge=1, le=5000),
        format: str = Query("csv", pattern="^(csv|xlsx)$"),
        category: Optional[str] = None,
        school_id: Optional[str] = None,
    ):
        """Export operacional (CSV/XLSX) dos casos prioritários — Fase 3A.1.

        Owner spec: secretarias vivem de planilha. Habilita imediatamente:
        Busca Ativa manual, reuniões pedagógicas, CRAS, Conselho Tutelar.

        Mesma engine `list_followup_cases` → consistência com endpoint JSON.
        Streaming attachment com `Content-Disposition`.
        """
        await AuthMiddleware.require_permission(
            db, 'nav-bolsa-familia-button', VIEW_ROLES
        )(request)

        payload = await list_followup_cases(
            db,
            academic_year=academic_year,
            mec_version=mec_version,
            severity_min=severity_min,
            limit=limit,
            category=category,
            school_id=school_id,
        )
        cases = payload.get("cases", [])
        # Resolve frequência consolidada (mês atual do caso) com engine canônica.
        # Owner exigiu coluna `frequência` na planilha.
        year_to_attendance: dict = {}
        async def _attendance_for(ay):
            if ay not in year_to_attendance:
                year_to_attendance[ay] = await db.attendance.find(
                    {"academic_year": ay}, {"_id": 0, "date": 1, "records": 1}
                ).to_list(50000)
            return year_to_attendance[ay]

        # Cache medical_days por (year, student) para reaproveitar
        med_cache: dict = {}
        async def _medical_for(ay, sid):
            key = (ay, sid)
            if key not in med_cache:
                m = await fetch_medical_days_for_students(db, [sid], ay)
                med_cache[key] = m.get(sid) or set()
            return med_cache[key]

        monthly_school_days_cache: dict = {}
        async def _school_days(ay):
            if ay not in monthly_school_days_cache:
                monthly_school_days_cache[ay] = await _calc_monthly_school_days(ay)
            return monthly_school_days_cache[ay]

        # Anota cada caso com frequência
        for c in cases:
            ay = c.get("academic_year")
            sid = c.get("student_id")
            month_str = str(c.get("month") or "")
            try:
                month_int = int(month_str)
            except (ValueError, TypeError):
                month_int = 0
            freq_str = ""
            if ay and sid and 1 <= month_int <= 12:
                att_docs = await _attendance_for(ay)
                med_set = await _medical_for(ay, sid)
                absences_map = compute_monthly_valid_absences(
                    att_docs, {sid: med_set}, {sid}
                )
                school_days = (await _school_days(ay)).get(month_int, 0)
                valid_abs = (absences_map.get(sid) or {}).get(month_int, 0)
                if school_days > 0:
                    freq_str = f"{round(((school_days - valid_abs) * 100) / school_days, 1)}%"
            c["frequency"] = freq_str

        # Render
        columns = [
            ("Aluno", "student_name"),
            ("Escola", "school_name"),
            ("Categoria MEC", "category"),
            ("Grupo MEC", "group_name"),
            ("Subcódigo", "reason_subcode"),
            ("Motivo", "reason_name"),
            ("Severidade", "severity_level"),
            ("Requer acompanhamento", "requires_followup"),
            ("Mês", "month"),
            ("Ano letivo", "academic_year"),
            ("Frequência", "frequency"),
            ("Observações", "notes"),
        ]
        filename_stem = f"bolsa_familia_busca_ativa_{academic_year or 'all'}"
        if category:
            filename_stem += f"_{category.lower()}"

        if format == "csv":
            import csv as _csv
            from io import StringIO
            buf = StringIO()
            writer = _csv.writer(buf, delimiter=";", quoting=_csv.QUOTE_MINIMAL)
            writer.writerow([h for h, _ in columns])
            for c in cases:
                writer.writerow([
                    "Sim" if c.get(k) is True else "Não" if c.get(k) is False else (c.get(k) or "")
                    for _, k in columns
                ])
            data = buf.getvalue().encode("utf-8-sig")  # BOM para abrir certo no Excel
            return StreamingResponse(
                io.BytesIO(data),
                media_type="text/csv; charset=utf-8",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename_stem}.csv"',
                    "X-Total-Cases": str(len(cases)),
                    "X-Stats-Version": NETWORK_STATS_VERSION,
                },
            )

        # xlsx
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = "Busca Ativa"
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1E40AF")
        for col_idx, (h, _) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for r_idx, c in enumerate(cases, start=2):
            for col_idx, (_, k) in enumerate(columns, start=1):
                val = c.get(k)
                if isinstance(val, bool):
                    val = "Sim" if val else "Não"
                ws.cell(row=r_idx, column=col_idx, value=val if val is not None else "")
        # Auto-width básico (limita 50)
        for col_idx, (h, _) in enumerate(columns, start=1):
            max_len = max(
                [len(str(c.get(columns[col_idx - 1][1]) or "")) for c in cases] + [len(h)]
            )
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 50)
        ws.freeze_panes = "A2"

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return StreamingResponse(
            out,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename_stem}.xlsx"',
                "X-Total-Cases": str(len(cases)),
                "X-Stats-Version": NETWORK_STATS_VERSION,
            },
        )

    @router.get("/bolsa-familia/students")
    async def list_bolsa_familia_students(
        request: Request,
        school_id: Optional[str] = Query(None, description="ID da escola; omitir para listar TODAS (apenas super_admin/admin/gerente/semed3)"),
        academic_year: Optional[int] = None,
        class_id: Optional[str] = Query(None, description="Filtra alunos por turma específica"),
    ):
        """Lista alunos com Bolsa Família de uma escola.

        Filtro opcional `class_id` (Fev/2026) — permite à secretaria
        focar em alunos de uma turma específica em vez do total da escola.

        Visão consolidada (Fev/2026): omitir `school_id` lista alunos BF
        de TODAS as escolas (apenas super_admin/admin/gerente/semed3).
        Esta visão é READ-ONLY — `can_edit=false` mesmo para EDIT_ROLES,
        já que o save é por-escola.
        """
        current_user = await AuthMiddleware.require_permission(
            db, 'nav-bolsa-familia-button', VIEW_ROLES
        )(request)

        all_schools_mode = not school_id
        if all_schools_mode and current_user.get('role') not in ALL_SCHOOLS_ROLES:
            raise HTTPException(
                status_code=403,
                detail="Apenas super_admin, admin, gerente ou semed3 podem consolidar todas as escolas."
            )

        if not academic_year:
            academic_year = datetime.now().year

        query = {
            "status": {"$in": ["active", "Ativo"]},
            "benefits": {"$in": ["Bolsa Família", "bolsa_familia", "Bolsa Familia"]}
        }
        if not all_schools_mode:
            query["school_id"] = school_id
        if class_id:
            query["class_id"] = class_id

        students = await db.students.find(
            query,
            {"_id": 0, "id": 1, "full_name": 1, "birth_date": 1, "nis": 1,
             "mother_name": 1, "class_id": 1, "school_id": 1,
             "inep_code": 1, "mother_phone": 1}
        ).sort("full_name", 1).collation({"locale": "pt", "strength": 1}).to_list(10000)

        class_ids = list(set(s.get("class_id") for s in students if s.get("class_id")))
        classes = await db.classes.find({"id": {"$in": class_ids}}, {"_id": 0, "id": 1, "name": 1, "grade_level": 1}).to_list(1000)
        class_map = {c["id"]: c for c in classes}

        # Mapa de school_id → school_name (somente em modo all-schools, evita query desnecessária)
        school_name_map: dict = {}
        if all_schools_mode:
            school_ids_in_result = list({s.get("school_id") for s in students if s.get("school_id")})
            if school_ids_in_result:
                schools_cursor = db.schools.find(
                    {"id": {"$in": school_ids_in_result}},
                    {"_id": 0, "id": 1, "name": 1},
                )
                async for sch in schools_cursor:
                    school_name_map[sch["id"]] = sch.get("name", "")

        # Buscar mantenedora
        mant = await get_mantenedora_cached(db)
        municipio_uf = ""
        if mant:
            mun = mant.get("municipio", "")
            uf = mant.get("uf", "")
            municipio_uf = f"{mun}/{uf}" if uf else mun

        # Calcular dias letivos por mês
        monthly_school_days = await _calc_monthly_school_days(academic_year)

        # Buscar tracking records (em modo all-schools, sem filtro por school_id)
        tracking_query = {"academic_year": academic_year}
        if not all_schools_mode:
            tracking_query["school_id"] = school_id
        bf_records = await db.bolsa_familia_tracking.find(
            tracking_query, {"_id": 0}
        ).to_list(50000)
        record_map = {}
        for r in bf_records:
            # Chave inclui school_id para evitar colisões em modo all-schools
            key = f"{r.get('school_id')}_{r['student_id']}_{r['month']}"
            record_map[key] = r

        # Calcular frequência mensal para todos os alunos.
        # FONTE ÚNICA DE VERDADE (Fev/2026 — Layer 1 LAYER do owner):
        # delega a contagem de FALTAS VÁLIDAS para
        # `services/attendance_utils.compute_monthly_valid_absences`, que
        # aplica as regras canônicas (atestado vence status; J não conta;
        # dependency_id ignorado). Substitui engine paralela que existia
        # aqui antes e que NÃO descontava atestados (bug crítico do BF).
        student_ids = [s["id"] for s in students]
        student_ids_set = set(student_ids)
        all_attendance = await db.attendance.find(
            {"academic_year": academic_year},
            {"_id": 0, "date": 1, "records": 1}
        ).to_list(50000)

        medical_days_by_student = await fetch_medical_days_for_students(
            db, student_ids, academic_year
        )
        student_absences = compute_monthly_valid_absences(
            all_attendance,
            medical_days_by_student,
            student_ids_set,
        )

        result = []
        for s in students:
            cls = class_map.get(s.get("class_id"), {})
            student_data = {
                "id": s["id"],
                "full_name": s["full_name"],
                "birth_date": s.get("birth_date", ""),
                "nis": s.get("nis", ""),
                "responsible": s.get("mother_name") or "",
                "contact": s.get("mother_phone") or "",
                "series": cls.get("grade_level") or cls.get("name", ""),
                "class_name": cls.get("name", ""),
                "inep_code": s.get("inep_code", ""),
                "school_id": s.get("school_id", ""),
                "school_name": school_name_map.get(s.get("school_id"), "") if all_schools_mode else "",
                "months": {}
            }

            for m in range(1, 13):
                key = f"{s.get('school_id')}_{s['id']}_{m}"
                rec = record_map.get(key, {})

                # Calcular frequência: ((dias_letivos - faltas) * 100) / dias_letivos
                school_days = monthly_school_days.get(m, 0)
                absences = (student_absences.get(s["id"]) or {}).get(m, 0)
                freq_pct = ""
                if school_days > 0:
                    freq_pct = f"{round(((school_days - absences) * 100) / school_days, 1)}%"

                # Conta dias de atestado nesse mês (transparência operacional — Fev/2026).
                # NÃO altera cálculo; apenas explica ao usuário por que freq pode
                # estar 100% mesmo com faltas no diário.
                medical_days_set = medical_days_by_student.get(s["id"]) or set()
                medical_days_in_month = sum(
                    1 for d in medical_days_set
                    if len(d) == 10 and d[5:7] == f"{m:02d}"
                )

                student_data["months"][str(m)] = {
                    "frequency": freq_pct,
                    "reason_id": rec.get("reason_id") or None,
                    "notes": rec.get("notes") or "",
                    "motive_legacy": rec.get("motive_legacy") or rec.get("motive") or "",
                    "school_days": school_days,
                    "absences": absences,
                    "medical_days_count": medical_days_in_month,
                    "has_medical_certificate": medical_days_in_month > 0,
                }

            result.append(student_data)

        return {
            "students": result,
            "total": len(result),
            "municipio_uf": municipio_uf,
            "monthly_school_days": monthly_school_days,
            "can_edit": False if all_schools_mode else (current_user['role'] in EDIT_ROLES),
            "all_schools_mode": all_schools_mode,
        }

    @router.put("/bolsa-familia/tracking")
    async def save_tracking(request: Request):
        """Salva dados de acompanhamento (reason_id/notes ou motivo legado).

        Schema novo: `reason_id` (referência a `attendance_frequency_reasons`) +
        `notes` (texto livre opcional).
        Mantém `motive_legacy` para compatibilidade retroativa (NÃO apaga).
        """
        current_user = await AuthMiddleware.require_permission(
            db, 'nav-bolsa-familia-button', EDIT_ROLES
        )(request)

        body = await request.json()
        student_id = body.get("student_id")
        school_id = body.get("school_id")
        month = body.get("month")
        academic_year = body.get("academic_year", datetime.now().year)
        reason_id = body.get("reason_id") or None
        notes = body.get("notes", "") or ""
        # Aceita motive legado APENAS se nenhum reason_id for fornecido (não criar novo legado).
        motive_legacy = body.get("motive_legacy") or body.get("motive") or ""

        if not student_id or not school_id or not month:
            raise HTTPException(status_code=400, detail="student_id, school_id e month são obrigatórios")

        # Validação: se reason_id, precisa existir
        if reason_id:
            exists = await db.attendance_frequency_reasons.find_one(
                {"id": reason_id, "active": True}, {"_id": 0, "id": 1}
            )
            if not exists:
                raise HTTPException(status_code=422, detail="reason_id inválido ou inativo")

        now = datetime.now(timezone.utc).isoformat()
        set_doc = {
            "student_id": student_id,
            "school_id": school_id,
            "month": str(month),
            "academic_year": academic_year,
            "reason_id": reason_id,
            "notes": notes,
            "updated_at": now,
            "saved_by_role": current_user.get('role'),
            "saved_by_user_id": current_user.get('id'),
        }
        # Só sobrescreve motive_legacy se foi explicitamente enviado.
        if motive_legacy:
            set_doc["motive_legacy"] = motive_legacy

        await db.bolsa_familia_tracking.update_one(
            {"student_id": student_id, "school_id": school_id, "month": str(month), "academic_year": academic_year},
            {"$set": set_doc},
            upsert=True
        )
        _cache_invalidate_all()
        return {"message": "Salvo com sucesso"}

    @router.put("/bolsa-familia/tracking/bulk")
    async def save_tracking_bulk(request: Request):
        """Salva em lote os motivos editados. Apenas EDIT_ROLES.

        Body: {"items": [{"student_id":..., "school_id":..., "month":...,
        "academic_year":..., "reason_id":..., "notes":...}, ...]}
        Retorna {ok, errors:[]}.
        """
        current_user = await AuthMiddleware.require_permission(
            db, 'nav-bolsa-familia-button', EDIT_ROLES
        )(request)
        body = await request.json()
        items = body.get("items") or []
        if not isinstance(items, list) or not items:
            raise HTTPException(status_code=400, detail="items é obrigatório (lista não vazia)")

        # Pré-carrega ids válidos uma vez (evita N queries)
        reason_ids = {it.get("reason_id") for it in items if it.get("reason_id")}
        valid_ids: set = set()
        if reason_ids:
            cursor = db.attendance_frequency_reasons.find(
                {"id": {"$in": list(reason_ids)}, "active": True},
                {"_id": 0, "id": 1},
            )
            async for r in cursor:
                valid_ids.add(r["id"])

        now = datetime.now(timezone.utc).isoformat()
        saved = 0
        errors: List[dict] = []
        for it in items:
            try:
                sid = it.get("student_id"); sch = it.get("school_id"); mo = it.get("month")
                ay = it.get("academic_year") or datetime.now().year
                rid = it.get("reason_id") or None
                if not (sid and sch and mo):
                    errors.append({"item": it, "error": "campos obrigatórios ausentes"}); continue
                if rid and rid not in valid_ids:
                    errors.append({"item": it, "error": "reason_id inválido ou inativo"}); continue
                set_doc = {
                    "student_id": sid,
                    "school_id": sch,
                    "month": str(mo),
                    "academic_year": ay,
                    "reason_id": rid,
                    "notes": it.get("notes", "") or "",
                    "updated_at": now,
                    "saved_by_role": current_user.get('role'),
                    "saved_by_user_id": current_user.get('id'),
                }
                motive_legacy = it.get("motive_legacy") or it.get("motive") or ""
                if motive_legacy:
                    set_doc["motive_legacy"] = motive_legacy
                await db.bolsa_familia_tracking.update_one(
                    {"student_id": sid, "school_id": sch, "month": str(mo), "academic_year": ay},
                    {"$set": set_doc},
                    upsert=True,
                )
                saved += 1
            except Exception as e:  # noqa: BLE001
                errors.append({"item": it, "error": str(e)})

        if saved > 0:
            _cache_invalidate_all()
        return {"saved": saved, "errors": errors}

    @router.get("/bolsa-familia/pdf/{school_id}")
    async def generate_bolsa_familia_pdf(
        school_id: str,
        request: Request,
        academic_year: Optional[int] = None,
        month_start: int = Query(2, description="Mês inicial"),
        month_end: int = Query(3, description="Mês final"),
        class_id: Optional[str] = Query(None, description="Filtra PDF por turma específica"),
    ):
        """Gera PDF de Acompanhamento de Frequência - Bolsa Família.

        Filtro opcional `class_id` (Fev/2026) — gera PDF segmentado por turma.
        """
        await AuthMiddleware.require_permission(
            db, 'nav-bolsa-familia-button', VIEW_ROLES
        )(request)

        if not academic_year:
            academic_year = datetime.now().year

        school = await db.schools.find_one({"id": school_id}, {"_id": 0})
        if not school:
            raise HTTPException(status_code=404, detail="Escola não encontrada")

        mant = await get_mantenedora_cached(db)
        municipio_uf = ""
        if mant:
            mun = mant.get("municipio", "")
            uf = mant.get("uf", "")
            municipio_uf = f"{mun}/{uf}" if uf else mun

        query = {
            "school_id": school_id,
            "status": {"$in": ["active", "Ativo"]},
            "benefits": {"$in": ["Bolsa Família", "bolsa_familia", "Bolsa Familia"]}
        }
        if class_id:
            query["class_id"] = class_id
        students = await db.students.find(
            query,
            {"_id": 0, "id": 1, "full_name": 1, "birth_date": 1, "nis": 1,
             "mother_name": 1, "class_id": 1, "inep_code": 1}
        ).sort("full_name", 1).collation({"locale": "pt", "strength": 1}).to_list(10000)

        if not students:
            detail = (
                "Nenhum aluno com Bolsa Família encontrado nesta turma"
                if class_id
                else "Nenhum aluno com Bolsa Família encontrado nesta escola"
            )
            raise HTTPException(status_code=404, detail=detail)

        class_ids = list(set(s.get("class_id") for s in students if s.get("class_id")))
        classes = await db.classes.find({"id": {"$in": class_ids}}, {"_id": 0, "id": 1, "name": 1, "grade_level": 1}).to_list(1000)
        class_map = {c["id"]: c for c in classes}

        bf_records = await db.bolsa_familia_tracking.find(
            {"school_id": school_id, "academic_year": academic_year},
            {"_id": 0}
        ).to_list(10000)
        record_map = {}
        reason_ids_used = set()
        for r in bf_records:
            key = f"{r['student_id']}_{r['month']}"
            record_map[key] = r
            if r.get("reason_id"):
                reason_ids_used.add(r["reason_id"])

        # Resolve nomes dos reasons usados em uma única query
        reason_name_map: dict = {}
        if reason_ids_used:
            cursor = db.attendance_frequency_reasons.find(
                {"id": {"$in": list(reason_ids_used)}},
                {"_id": 0, "id": 1, "name": 1, "mec_subcode": 1},
            )
            async for r in cursor:
                reason_name_map[r["id"]] = f"{r.get('mec_subcode', '')} - {r.get('name', '')}".strip(" -")

        monthly_school_days = await _calc_monthly_school_days(academic_year)

        # Buscar frequência — usa engine canônica (Fev/2026 Layer 1 fix).
        # Atestado médico vence falta; J nunca conta; dependency_id ignorado.
        student_ids = [s["id"] for s in students]
        all_attendance = await db.attendance.find(
            {"academic_year": academic_year},
            {"_id": 0, "date": 1, "records": 1}
        ).to_list(50000)

        medical_days_by_student = await fetch_medical_days_for_students(
            db, student_ids, academic_year
        )
        student_absences = compute_monthly_valid_absences(
            all_attendance,
            medical_days_by_student,
            set(student_ids),
        )

        secretario = school.get("secretario_escolar") or ""
        months_range = list(range(month_start, month_end + 1))

        # [Fev/2026] Validação digital: a assinatura digital aparece no PDF SOMENTE
        # se TODOS os meses do range tiverem AO MENOS UM tracking salvo por
        # `saved_by_role='secretario'` (autenticidade do registro pela escola).
        digital_signer_name = ""
        digital_signature_valid = False
        try:
            tracking_secretary = await db.bolsa_familia_tracking.find(
                {
                    "school_id": school_id,
                    "academic_year": academic_year,
                    "saved_by_role": "secretario",
                    "month": {"$in": [str(m) for m in months_range]},
                },
                {"_id": 0, "month": 1, "saved_by_user_id": 1},
            ).to_list(50000)

            months_signed = {str(t.get("month")) for t in tracking_secretary}
            all_signed = all(str(m) in months_signed for m in months_range)
            if all_signed and tracking_secretary:
                # Pega o nome do servidor que assinou (último signatário do range)
                signer_id = tracking_secretary[-1].get("saved_by_user_id")
                if signer_id:
                    signer = await db.staff.find_one(
                        {"id": signer_id}, {"_id": 0, "nome": 1}
                    ) or await db.users.find_one(
                        {"id": signer_id}, {"_id": 0, "name": 1, "full_name": 1}
                    )
                    if signer:
                        digital_signer_name = (
                            signer.get("nome") or signer.get("full_name") or signer.get("name") or ""
                        )
                if not digital_signer_name and secretario:
                    digital_signer_name = secretario
                digital_signature_valid = bool(digital_signer_name)
        except Exception as e:  # noqa: BLE001
            logger.warning(f"Falha ao validar assinatura digital BF: {e}")

        try:
            pdf_buffer = _generate_bf_pdf(
                school=school, students=students, class_map=class_map,
                record_map=record_map, months_range=months_range,
                academic_year=academic_year, secretario=secretario,
                municipio_uf=municipio_uf, monthly_school_days=monthly_school_days,
                student_absences=student_absences,
                digital_signer_name=digital_signer_name,
                digital_signature_valid=digital_signature_valid,
                reason_name_map=reason_name_map,
            )
        except Exception as e:
            logger.error(f"Erro ao gerar PDF Bolsa Família: {e}")
            raise HTTPException(status_code=500, detail=f"Erro ao gerar PDF: {str(e)}")

        filename = f"bolsa_familia_{school.get('name', 'escola').replace(' ', '_')}_{academic_year}.pdf"
        return StreamingResponse(
            pdf_buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"inline; filename={filename}"}
        )

    def _generate_bf_pdf(school, students, class_map, record_map, months_range, academic_year, secretario, municipio_uf, monthly_school_days, student_absences, digital_signer_name="", digital_signature_valid=False, reason_name_map=None):
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm, cm
        from reportlab.lib.enums import TA_LEFT, TA_CENTER
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1.5*cm, bottomMargin=1.5*cm, leftMargin=1.5*cm, rightMargin=1.5*cm)
        elements = []

        title_style = ParagraphStyle('BFTitle', fontSize=12, fontName='Helvetica-Bold', alignment=TA_CENTER, leading=14, spaceAfter=4)
        info_style = ParagraphStyle('BFInfo', fontSize=7, leading=9, alignment=TA_LEFT)
        cell_style = ParagraphStyle('BFCell', fontSize=6.5, leading=8, alignment=TA_LEFT)
        cell_center = ParagraphStyle('BFCellCenter', fontSize=6.5, leading=8, alignment=TA_CENTER)
        sign_style = ParagraphStyle('BFSign', fontSize=8, alignment=TA_CENTER, leading=10)

        elements.append(Paragraph("Acompanhamento de Frequência Escolar", title_style))
        elements.append(Spacer(1, 4))

        school_name = school.get('name', '')
        inep = school.get('inep_code', '')

        school_data = [
            [Paragraph(f"<b>Nome da Escola:</b> {school_name}", info_style),
             Paragraph(f"<b>Código INEP:</b> {inep}", info_style),
             Paragraph(f"<b>Município/UF:</b> {municipio_uf}", info_style)]
        ]
        school_table = Table(school_data, colWidths=[8*cm, 4.5*cm, 4.5*cm])
        school_table.setStyle(TableStyle([
            ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(school_table)
        elements.append(Spacer(1, 6))

        for student in students:
            cls = class_map.get(student.get("class_id"), {})
            serie = cls.get("grade_level") or cls.get("name", "")
            birth = student.get("birth_date", "")
            if isinstance(birth, str) and "-" in birth:
                try:
                    bd = datetime.strptime(birth.split("T")[0], "%Y-%m-%d")
                    birth = bd.strftime("%d/%m/%Y")
                except:
                    pass
            nis = student.get("nis", "")
            responsible = student.get("mother_name") or ""
            student_inep = student.get("inep_code", "")

            student_header = [
                [Paragraph(f"<b>Nome do Estudante:</b> {student['full_name']}", info_style),
                 Paragraph(f"<b>Dt. Nasc.:</b> {birth}", info_style),
                 Paragraph(f"<b>NIS:</b> {nis}", info_style)],
                [Paragraph(f"<b>Responsável familiar:</b> {responsible}", info_style),
                 Paragraph(f"<b>Código INEP:</b> {student_inep}", info_style),
                 Paragraph(f"<b>Série:</b> {serie}", info_style)]
            ]
            header_table = Table(student_header, colWidths=[8*cm, 4.5*cm, 4.5*cm])
            header_table.setStyle(TableStyle([
                ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
                ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('BACKGROUND', (0, 0), (-1, -1), colors.Color(0.95, 0.95, 0.95)),
            ]))
            elements.append(header_table)

            freq_header = [Paragraph("<b>Mês</b>", cell_center),
                          Paragraph("<b>Frequência</b>", cell_center),
                          Paragraph("<b>Motivo</b>", cell_center)]
            freq_data = [freq_header]

            for m in months_range:
                key = f"{student['id']}_{m}"
                rec = record_map.get(key, {})

                # Novo schema: reason_id + notes; fallback motive_legacy
                rid = rec.get("reason_id")
                reason_name = (reason_name_map or {}).get(rid, "") if rid else ""
                notes_val = rec.get("notes") or ""
                legacy_val = rec.get("motive_legacy") or rec.get("motive") or ""
                if reason_name:
                    motive_val = reason_name + (f" — {notes_val}" if notes_val else "")
                elif legacy_val:
                    motive_val = legacy_val
                else:
                    motive_val = notes_val

                school_days = monthly_school_days.get(m, 0)
                absences = (student_absences.get(student["id"]) or {}).get(m, 0)
                freq_pct = ""
                if school_days > 0:
                    freq_pct = f"{round(((school_days - absences) * 100) / school_days, 1)}%"

                freq_data.append([
                    Paragraph(MESES_PT.get(m, str(m)), cell_center),
                    Paragraph(freq_pct, cell_center),
                    Paragraph(motive_val, cell_style)
                ])

            freq_table = Table(freq_data, colWidths=[3.5*cm, 3.5*cm, 10*cm])
            freq_table.setStyle(TableStyle([
                ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
                ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.85, 0.85, 0.85)),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ]))
            elements.append(freq_table)
            elements.append(Spacer(1, 8))

        elements.append(Spacer(1, 20))
        if digital_signature_valid and digital_signer_name:
            # Assinatura digital: aparece SOMENTE quando todos os meses do range
            # foram salvos por um secretário da escola.
            sign_paragraph = f"Documento validado digitalmente por <b>{digital_signer_name}</b>"
            sub_paragraph = f"Secretário(a) da {school.get('name', '')}"
            elements.append(Paragraph(sign_paragraph, sign_style))
            elements.append(Paragraph(sub_paragraph, sign_style))
        else:
            elements.append(Paragraph("_" * 60, sign_style))
            elements.append(Paragraph("Assinatura do Responsável pelas Informações na Escola", sign_style))

        doc.build(elements)
        buffer.seek(0)
        return buffer

    return router
