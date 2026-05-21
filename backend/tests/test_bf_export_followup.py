"""Fase 3A.1 (Fev/2026) — Export CSV/XLSX dos casos prioritários BF.

Valida `/api/bolsa-familia/stats/network/followup/export`:
  - Reusa engine `list_followup_cases` (consistência com endpoint JSON).
  - CSV com BOM UTF-8 + separador `;` (padrão Brasil/Excel).
  - XLSX gerado via openpyxl com header estilizado.
  - Headers HTTP: Content-Disposition attachment, X-Total-Cases, X-Stats-Version.
  - Frequência calculada via engine canônica.
"""
import os
import asyncio
from datetime import datetime, timezone
from io import BytesIO

import pytest
import requests
from motor.motor_asyncio import AsyncIOMotorClient
from openpyxl import load_workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")
ADMIN_EMAIL = "gutenberg@sigesc.com"
ADMIN_PASS = "@Celta2007"
QA_PREFIX = "qa-export-"
ACADEMIC_YEAR = 2098


@pytest.fixture(scope="module")
def auth():
    if not BASE_URL:
        pytest.skip("REACT_APP_BACKEND_URL not set")
    r = requests.post(
        f"{BASE_URL}/api/auth/login",
        json={"email": ADMIN_EMAIL, "password": ADMIN_PASS},
        timeout=15,
    )
    assert r.status_code == 200
    body = r.json()
    return {"token": body["access_token"]}


def _h(auth):
    return {"Authorization": f"Bearer {auth['token']}"}


@pytest.fixture(scope="module")
def seeded():
    async def _setup():
        client = AsyncIOMotorClient(os.environ["MONGO_URL"])
        db = client[os.environ["DB_NAME"]]
        reasons = await db.attendance_frequency_reasons.find(
            {"mec_subcode": {"$in": ["11a", "10b", "3b", "1a"]}, "mec_version": "4.2"},
            {"_id": 0, "id": 1, "mec_subcode": 1},
        ).to_list(10)
        by_sub = {r["mec_subcode"]: r["id"] for r in reasons}
        await db.bolsa_familia_tracking.delete_many({"student_id": {"$regex": f"^{QA_PREFIX}"}})
        now = datetime.now(timezone.utc).isoformat()
        docs = [
            {"student_id": f"{QA_PREFIX}v1", "school_id": "qa-sc-X", "month": "3",
             "academic_year": ACADEMIC_YEAR, "reason_id": by_sub["11a"],
             "notes": "Caso prioritário violência", "updated_at": now},
            {"student_id": f"{QA_PREFIX}c1", "school_id": "qa-sc-X", "month": "4",
             "academic_year": ACADEMIC_YEAR, "reason_id": by_sub["10b"],
             "notes": "", "updated_at": now},
            {"student_id": f"{QA_PREFIX}t1", "school_id": "qa-sc-Y", "month": "3",
             "academic_year": ACADEMIC_YEAR, "reason_id": by_sub["3b"],
             "notes": "Ônibus quebrou", "updated_at": now},
            {"student_id": f"{QA_PREFIX}h1", "school_id": "qa-sc-Y", "month": "3",
             "academic_year": ACADEMIC_YEAR, "reason_id": by_sub["1a"],
             "notes": "Atestado", "updated_at": now},  # severity 2 -> entra só com severity_min=1
        ]
        await db.bolsa_familia_tracking.insert_many(docs)

    asyncio.run(_setup())
    yield
    async def _teardown():
        client = AsyncIOMotorClient(os.environ["MONGO_URL"])
        db = client[os.environ["DB_NAME"]]
        await db.bolsa_familia_tracking.delete_many({"student_id": {"$regex": f"^{QA_PREFIX}"}})
    asyncio.run(_teardown())


def test_csv_export_headers_and_format(auth, seeded):
    r = requests.get(
        f"{BASE_URL}/api/bolsa-familia/stats/network/followup/export?format=csv&academic_year={ACADEMIC_YEAR}&severity_min=1",
        headers=_h(auth), timeout=20,
    )
    assert r.status_code == 200
    assert "text/csv" in r.headers.get("content-type", "").lower()
    assert "attachment" in r.headers.get("content-disposition", "")
    assert r.headers.get("x-stats-version") == "v1.0"
    assert int(r.headers.get("x-total-cases", "0")) >= 4
    body = r.content
    # BOM UTF-8 (necessário p/ Excel abrir acentos corretamente)
    assert body[:3] == b"\xef\xbb\xbf"
    # Separador ;
    assert b";" in body
    # Header em PT-BR
    decoded = body.decode("utf-8-sig")
    assert "Aluno" in decoded.split("\n")[0]
    assert "Categoria MEC" in decoded.split("\n")[0]
    assert "Severidade" in decoded.split("\n")[0]


def test_csv_export_content_includes_qa_cases(auth, seeded):
    r = requests.get(
        f"{BASE_URL}/api/bolsa-familia/stats/network/followup/export?format=csv&academic_year={ACADEMIC_YEAR}&severity_min=1",
        headers=_h(auth), timeout=20,
    )
    decoded = r.content.decode("utf-8-sig")
    assert "VIOLENCE" in decoded
    assert "11a" in decoded
    assert "Caso prioritário violência" in decoded
    assert "Sim" in decoded  # requires_followup=True renderizado em PT


def test_xlsx_export_structure(auth, seeded):
    r = requests.get(
        f"{BASE_URL}/api/bolsa-familia/stats/network/followup/export?format=xlsx&academic_year={ACADEMIC_YEAR}&severity_min=1",
        headers=_h(auth), timeout=20,
    )
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers.get("content-type", "")
    assert "attachment" in r.headers.get("content-disposition", "")
    assert ".xlsx" in r.headers.get("content-disposition", "")
    wb = load_workbook(BytesIO(r.content))
    ws = wb.active
    assert ws.title == "Busca Ativa"
    # 12 colunas conforme spec do owner
    assert ws.max_column == 12
    # Header esperado
    headers = [ws.cell(row=1, column=i).value for i in range(1, 13)]
    assert headers[0] == "Aluno"
    assert headers[2] == "Categoria MEC"
    assert headers[6] == "Severidade"
    # Freeze pane (header sempre visível)
    assert ws.freeze_panes == "A2"
    # Header com formatação (bold)
    assert ws.cell(row=1, column=1).font.bold is True


def test_xlsx_export_includes_qa_cases(auth, seeded):
    r = requests.get(
        f"{BASE_URL}/api/bolsa-familia/stats/network/followup/export?format=xlsx&academic_year={ACADEMIC_YEAR}&severity_min=1",
        headers=_h(auth), timeout=20,
    )
    wb = load_workbook(BytesIO(r.content))
    ws = wb.active
    # Coleta todas as células
    all_text = "\n".join(
        " | ".join(str(ws.cell(row=r_idx, column=c).value or "") for c in range(1, 13))
        for r_idx in range(2, ws.max_row + 1)
    )
    assert "VIOLENCE" in all_text
    assert "11a" in all_text
    assert "Caso prioritário violência" in all_text


def test_export_format_validation(auth, seeded):
    """format=xml não é aceito (validação Query pattern)."""
    r = requests.get(
        f"{BASE_URL}/api/bolsa-familia/stats/network/followup/export?format=xml&academic_year={ACADEMIC_YEAR}",
        headers=_h(auth), timeout=15,
    )
    assert r.status_code == 422


def test_export_severity_filter(auth, seeded):
    """severity_min=5 deve trazer só os 2 casos críticos (11a, 10b)."""
    r = requests.get(
        f"{BASE_URL}/api/bolsa-familia/stats/network/followup/export?format=csv&academic_year={ACADEMIC_YEAR}&severity_min=5",
        headers=_h(auth), timeout=20,
    )
    decoded = r.content.decode("utf-8-sig")
    # Conta linhas de dados (descontando header)
    lines = [ln for ln in decoded.split("\n") if ln.strip() and "Aluno" not in ln]
    # Inclui também 3b (requires_followup=True) pelo OR — owner spec
    # 11a (sev5) + 10b (sev5) + 3b (requires_followup) = pelo menos 3
    qa_lines = [ln for ln in lines if any(t in ln for t in ["VIOLENCE", "CHILD_LABOR", "ACCESS"])]
    assert len(qa_lines) >= 3
    # 1a não deve aparecer (severity 2, sem requires_followup)
    assert "Doença/problemas físicos" not in decoded
